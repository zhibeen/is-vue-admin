from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ExcelService:
    def generate_contracts_excel(self, contracts):
        wb = Workbook()
        # Remove default sheet
        default_ws = wb.active
        wb.remove(default_ws)
        
        # Group by Company -> Supplier
        # Sort contracts by Company first to ensure consistent grouping order
        contracts.sort(key=lambda x: (x.company.short_name if x.company and x.company.short_name else "Unknown"))
        
        grouped = {}
        for c in contracts:
            comp_name = c.company.short_name if c.company and c.company.short_name else (c.company.legal_name if c.company else "Unknown")
            if comp_name not in grouped:
                grouped[comp_name] = []
            grouped[comp_name].append(c)
            
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Updated Headers
        headers = ["供应商", "合同编号", "创建日期", "业务状态", "SKU", "产品名称", "数量", "单价", "行总价", "合同总额", "已付金额", "未付余额", "付款条款"]
        
        for comp_name, comp_contracts in grouped.items():
            # Sheet name limit 31 chars, avoid invalid chars
            safe_sheet_name = "".join([c for c in comp_name if c not in ":\\/?*[]"])[:30]
            ws = wb.create_sheet(title=safe_sheet_name) 
            
            # Write Header
            ws.append(headers)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            
            # Sort contracts by Supplier -> Date
            comp_contracts.sort(key=lambda x: (x.supplier.name, x.created_at))
            
            current_row = 2
            currency_summary = {} # {'CNY': 0, 'USD': 0}
            
            for contract in comp_contracts:
                # Accumulate Summary
                curr = contract.currency or 'CNY'
                if curr not in currency_summary:
                    currency_summary[curr] = 0
                currency_summary[curr] += float(contract.total_amount)

                items = contract.items
                # Snapshot Logic for Supplier Name
                supplier_name = contract.supplier.name
                if contract.supplier_snapshot and isinstance(contract.supplier_snapshot, dict) and 'name' in contract.supplier_snapshot:
                    # Use snapshot name, maybe append a marker or just use it
                    supplier_name = contract.supplier_snapshot['name']

                # Financials
                paid_amt = float(contract.paid_amount)
                unpaid_amt = float(contract.total_amount) - paid_amt

                # Smart resolve payment term name
                term_name = ""
                if contract.supplier_snapshot and isinstance(contract.supplier_snapshot, dict) and 'payment_terms' in contract.supplier_snapshot:
                     term_name = contract.supplier_snapshot['payment_terms']
                elif contract.payment_term:
                    term_name = contract.payment_term.name
                elif contract.payment_terms:
                    term_name = contract.payment_terms

                if not items:
                    # Handle empty contract
                    row_data = [
                        supplier_name,
                        contract.contract_no,
                        contract.created_at.strftime('%Y-%m-%d'),
                        contract.status,
                        "", "", 0, 0, 0,
                        float(contract.total_amount),
                        paid_amt,
                        unpaid_amt,
                        term_name
                    ]
                    ws.append(row_data)
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.border = border
                        cell.alignment = center_align
                    current_row += 1
                    continue
                    
                start_row = current_row
                
                for item in items:
                    row_data = [
                        supplier_name,
                        contract.contract_no,
                        contract.created_at.strftime('%Y-%m-%d'),
                        contract.status,
                        item.product.sku if item.product else "",
                        item.product.name if item.product else "",
                        float(item.confirmed_qty),
                        float(item.unit_price),
                        float(item.total_price),
                        float(contract.total_amount),
                        paid_amt,
                        unpaid_amt,
                        term_name
                    ]
                    
                    ws.append(row_data)
                    
                    # Style each cell
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.border = border
                        if col_num in [7, 8, 9, 10, 11, 12]: # Numeric columns
                             cell.number_format = '#,##0.00'
                        else:
                             cell.alignment = center_align

                    current_row += 1
                
                end_row = current_row - 1
                
                # Merge cells for contract-level info
                if end_row > start_row:
                    # Merge Supplier, Contract No, Date, Status, Total Amount, Paid, Unpaid, Terms
                    # Cols: 1, 2, 3, 4, 10, 11, 12, 13
                    merge_cols = [1, 2, 3, 4, 10, 11, 12, 13]
                    for col in merge_cols:
                        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=col, end_column=col)
                        # Re-apply alignment to top-left cell of merge range
                        cell = ws.cell(row=start_row, column=col)
                        cell.alignment = center_align

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = min(adjusted_width, 50) 

            # Append Summary
            current_row += 2
            ws.cell(row=current_row, column=1, value="资金汇总 (按币种)").font = Font(bold=True)
            current_row += 1
            
            for curr, amt in currency_summary.items():
                ws.cell(row=current_row, column=1, value=curr).alignment = center_align
                ws.cell(row=current_row, column=2, value=amt).number_format = '#,##0.00'
                ws.cell(row=current_row, column=2).font = Font(bold=True)
                current_row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

excel_service = ExcelService()
